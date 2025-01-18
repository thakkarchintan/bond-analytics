import openpyxl
from datetime import datetime
import os
from bs4 import BeautifulSoup
from playwright.sync_api import sync_playwright
import time
from playwright_stealth import stealth_sync
import random
import os
import subprocess
import logging
from datetime import datetime


logging.basicConfig(level=logging.INFO)


def has_function_run_today(log_file="last_run_date.txt"):
    """Check if the function has already run today."""
    if os.path.exists(log_file):
        with open(log_file, "r") as file:
            last_run_date = file.read().strip()
        # Compare the last run date with today's date
        if last_run_date == datetime.today().strftime("%Y-%m-%d"):
            return True
    return False

def update_last_run_date(log_file="last_run_date.txt"):
    """Update the last run date to today's date."""
    with open(log_file, "w") as file:
        file.write(datetime.today().strftime("%Y-%m-%d"))

def move_mouse_randomly(page):
    """
    Simulates human-like mouse movements by moving to a random position on the page.
    """
    x, y = random.randint(100, 500), random.randint(100, 500)
    page.mouse.move(x, y)
    
def has_function_run_today(log_file="last_run_date.txt"):
    """Check if the function has already run today."""
    if os.path.exists(log_file):
        with open(log_file, "r") as file:
            last_run_date = file.read().strip()
        # Compare the last run date with today's date
        if last_run_date == datetime.today().strftime("%Y-%m-%d"):
            return True
    return False

def update_last_run_date(log_file="last_run_date.txt"):
    """Update the last run date to today's date."""
    with open(log_file, "w") as file:
        file.write(datetime.today().strftime("%Y-%m-%d"))

def update_data(url, tool, append_dates=False , retries=5):
    with sync_playwright() as p:
        # Launch browser with additional arguments and settings
        browser = p.chromium.launch(headless=True, args=[
        '--disable-blink-features=AutomationControlled',
        '--ignore-certificate-errors',
        '--no-sandbox',
        '--disable-dev-shm-usage'
        ])

        # Create a new browser context with a custom user agent
        context = browser.new_context(
            user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36",
            locale="en-US",
            timezone_id="America/New_York",
            geolocation={"latitude": 40.7128, "longitude": -74.0060},  # Example for New York
            permissions=["geolocation"],
            viewport={"width": 1280, "height": 720},  # Correct viewport format
            ignore_https_errors=True  # Ensures SSL certificate issues are ignored
        )
        
        page = context.new_page()
        
        stealth_sync(page)

        # try:
        #     # Navigate to the URL
        #     page.goto(url , timeout=60000)
        #     # Wait for the table to load dynamically
        #     page.wait_for_selector(".freeze-column-w-1", state="visible", timeout=60000)  # Wait for the specific element
        #     print(f"Page loaded successfully for {tool}")
        # except Exception as e:
        #     print(f"Error while loading page: {e}")
        #     browser.close()
        #     return
        
        for attempt in range(retries):
            try:
                # Attempt to load the page
                print(f"Attempting to load {url} (Attempt {attempt + 1})")
                page.goto(url)
                # Simulate user interaction to avoid detection
                move_mouse_randomly(page)
                # Wait for the specific selector indicating the page is loaded
                page.wait_for_selector(".freeze-column-w-1")
                print(f"Successfully loaded {url} on attempt {attempt + 1}")
                break
            except Exception as e:
                print(f"Error loading {url} on attempt {attempt + 1}: {e}")
                if attempt < retries - 1:
                    print("Retrying...")
                    time.sleep(5)  # Wait before retrying
                else:
                    print("Max retries reached. Could not load the page.")
                    return False

        # Parse the page content
        html = page.content()
        soup = BeautifulSoup(html, "html.parser")
        
        logging.info("Script started")
        
        # Extract the table
        table = soup.find("table", class_="freeze-column-w-1")
        data = []

        if table:
            rows = table.find("tbody").find_all("tr")
            for row in rows:
                try:
                    # Extract Date
                    date = row.find("time").get("datetime")

                    # Extract Price
                    price_td = row.find("td", class_="datatable_cell--align-end__qgxDQ")
                    price = price_td.get_text(strip=True) if price_td else None

                    # Append data if price is found
                    if price:
                        data.append(
                            {"Date": date, "Price": float(price.replace(",", ""))}  # Convert to float
                        )
                except Exception as e:
                    print(f"Error extracting row data for {tool}: {e}")
        else:
            print(f"Table not found for {tool}")

        # Reverse the data to ensure chronological order
        data.reverse()

        # Open the Excel workbook
        file_path = "Final.xlsx"
        wb = openpyxl.load_workbook(file_path)
        sheet = wb.active

        today_date = datetime.today().date()  # Today's date

        # Process data and update Excel
        for entry in data:
            try:
                entry_date = datetime.strptime(entry["Date"], "%b %d, %Y").date()
            except ValueError:
                print(f"Failed to parse date: {entry['Date']}")
                continue

            # Only process entries with dates less than today
            if entry_date >= today_date:
                continue

            # Append dates if this is the first tool and append_dates is True
            if append_dates:
                # Check if the date already exists
                date_exists = False
                for row in range(2, sheet.max_row + 1):
                    excel_date = sheet.cell(row=row, column=1).value
                    if isinstance(excel_date, datetime):
                        excel_date = excel_date.date()
                    elif isinstance(excel_date, str):
                        excel_date = datetime.strptime(excel_date, "%d-%b-%Y").date()

                    if excel_date == entry_date:
                        date_exists = True
                        break

                # Append the date if it does not exist
                if not date_exists:
                    new_row = sheet.max_row + 1
                    sheet.cell(row=new_row, column=1).value = entry_date.strftime("%d-%b-%Y")

            # Update or add the price for the tool
            tool_column_index = None
            for col_idx, cell in enumerate(sheet[1], start=1):
                if cell.value == tool:
                    tool_column_index = col_idx
                    break

            if tool_column_index is None:
                # Create a new column for the tool
                tool_column_index = sheet.max_column + 1
                sheet.cell(row=1, column=tool_column_index).value = tool

            # Update the price in the corresponding row
            for row in range(2, sheet.max_row + 1):
                excel_date = sheet.cell(row=row, column=1).value
                if isinstance(excel_date, datetime):
                    excel_date = excel_date.date()
                elif isinstance(excel_date, str):
                    excel_date = datetime.strptime(excel_date, "%d-%b-%Y").date()

                if excel_date == entry_date:
                    sheet.cell(row=row, column=tool_column_index).value = entry["Price"]
                    break
            else:
                # Append a new row if the date is missing
                new_row = sheet.max_row + 1
                sheet.cell(row=new_row, column=1).value = entry_date.strftime("%d-%b-%Y")
                sheet.cell(row=new_row, column=tool_column_index).value = entry["Price"]

        # Save the workbook
        try:
            wb.save(file_path)
            print(f"Excel updated successfully for {tool}!")
        except PermissionError:
            print(f"Error: Unable to save the file '{file_path}'. Please ensure it is not open in another application.")
        
        # Close the browser
        browser.close()


def automate_update_excel() :
    url_data = [
        {'US2Y': 'https://in.investing.com/rates-bonds/u.s.-2-year-bond-yield-historical-data'},
        {'US5Y': 'https://in.investing.com/rates-bonds/u.s.-5-year-bond-yield-historical-data'},
        {'US10Y': 'https://in.investing.com/rates-bonds/u.s.-10-year-bond-yield-historical-data'},
        {'US30Y': 'https://in.investing.com/rates-bonds/u.s.-30-year-bond-yield-historical-data'},
        {'FGBSY': 'https://in.investing.com/rates-bonds/germany-2-year-bond-yield-historical-data'},
        {'FGBMY': 'https://in.investing.com/rates-bonds/germany-5-year-bond-yield-historical-data'},
        {'FGBLY': 'https://in.investing.com/rates-bonds/germany-10-year-bond-yield-historical-data'},
        {'FGBXY': 'https://in.investing.com/rates-bonds/germany-30-year-bond-yield-historical-data'},
        {'CAD2Y': 'https://www.investing.com/rates-bonds/canada-2-year-bond-yield-historical-data'},
        {'CAD3Y': 'https://www.investing.com/rates-bonds/canada-3-year-bond-yield-historical-data'},
        {'CAD5Y': 'https://www.investing.com/rates-bonds/canada-5-year-bond-yield-historical-data'},
        {'CAD10Y': 'https://www.investing.com/rates-bonds/canada-10-year-bond-yield-historical-data'},
        {'UK10Y': 'https://in.investing.com/rates-bonds/uk-10-year-bond-yield-historical-data'},
        {'AUS10Y': 'https://in.investing.com/rates-bonds/australia-10-year-bond-yield-historical-data'},
        {'FBTPY': 'https://www.investing.com/rates-bonds/italy-10-year-bond-yield-historical-data'},
        {'FBTSY': 'https://www.investing.com/rates-bonds/italy-2-year-bond-yield-historical-data'},
        {'FOATY': 'https://www.investing.com/rates-bonds/france-10-year-bond-yield-historical-data'},
    ]

    # Run the first tool with append_dates=True
    for i, entry in enumerate(url_data):
        for tool, url in entry.items():
            update_data(url, tool, append_dates=(i == 0))
    
    update_last_run_date()
    
    
if not has_function_run_today():
    automate_update_excel()
    
repo_path = os.path.dirname(os.path.abspath(__file__))

if os.path.exists(os.path.join(repo_path, ".git")):
    print(f"The repository path is: {repo_path}")
else:
    print(f"'{repo_path}' is not a valid Git repository.")

    
try:
    file_path = "Final.xlsx"
    today = datetime.now()
    formatted_date = today.strftime("%d-%m-%Y")
    # Add all changes
    subprocess.run(["git", "add", file_path], cwd=repo_path, check=True)

    # Commit changes
    commit_message = f"Excel data updated for date{formatted_date}"
    subprocess.run(["git", "commit", "-m", commit_message], cwd=repo_path, check=True)

    # Push changes
    subprocess.run(["git", "push"], cwd=repo_path, check=True)

    print("Changes added, committed, and pushed successfully!")
except subprocess.CalledProcessError as e:
    print(f"Error: {e}")